import csv
import io
import json
import re
from pathlib import Path

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA_FILE = ROOT / "data" / "inquiries.json"
HEADERS = ["受付日", "問い合わせ内容", "問い合わせ種類", "回答内容", "対象カテゴリ", "他部署へのエスカレーション有無", "問い合わせ言語", "該当データを回答できたか"]
app = FastAPI(title="問い合わせナレッジ検索")

def records():
    return json.loads(DATA_FILE.read_text(encoding="utf-8"))

def grams(value):
    text = re.sub(r"[\s　、。,.!?！？・（）()]", "", str(value).lower())
    return {text[i:i + 2] for i in range(max(len(text) - 1, 1))} if text else set()

def similarity(query, text):
    left, right = grams(query), grams(text)
    return len(left & right) / len(left | right) if left and right else 0

def score(row, query):
    if not query.strip():
        return 0
    return min(100, round(100 * (similarity(query, row["問い合わせ内容"]) * .65 + similarity(query, row["回答内容"]) * .25 + similarity(query, row["対象カテゴリ"]) * .10)))

@app.get("/api/options")
def options():
    data = records()
    return {field: sorted({row[field] for row in data if row[field]}) for field in ["問い合わせ種類", "対象カテゴリ", "問い合わせ言語"]}

@app.get("/api/search")
def search(q: str = "", inquiry_type: str = "", category: str = "", language: str = "", answerable: str = ""):
    data = records()
    filtered = [row for row in data if (not inquiry_type or row["問い合わせ種類"] == inquiry_type) and (not category or row["対象カテゴリ"] == category) and (not language or row["問い合わせ言語"] == language) and (not answerable or row["該当データを回答できたか"] == answerable)]
    output = [{**row, "similarity": score(row, q)} for row in filtered]
    output.sort(key=lambda row: (row["similarity"], row["受付日"]), reverse=True)
    return {"total": len(output), "results": output[:10]}

@app.post("/api/import")
async def import_excel(file: UploadFile = File(...)):
    raw, suffix = await file.read(), Path(file.filename or "").suffix.lower()
    try:
        if suffix == ".csv":
            rows = list(csv.reader(io.StringIO(raw.decode("utf-8-sig"))))
        elif suffix in {".xlsx", ".xlsm"}:
            rows = list(load_workbook(io.BytesIO(raw), read_only=True, data_only=True).active.iter_rows(values_only=True))
        else:
            raise HTTPException(400, "CSV または XLSX ファイルを選択してください。")
        headers = [str(value or "").strip() for value in rows[0]]
        missing = [field for field in HEADERS if field not in headers]
        if missing:
            raise HTTPException(400, "必要な列がありません: " + "、".join(missing))
        data = [{field: str(row[headers.index(field)] or "").strip() for field in HEADERS} for row in rows[1:] if any(value is not None and str(value).strip() for value in row)]
        DATA_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"message": f"{len(data)}件を読み込みました。"}
    except HTTPException:
        raise
    except Exception as error:
        raise HTTPException(400, f"読み込みに失敗しました: {error}")

app.mount("/static", StaticFiles(directory=ROOT / "app" / "static"), name="static")

@app.get("/")
def index():
    return FileResponse(ROOT / "app" / "static" / "index.html")

